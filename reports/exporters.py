"""Render a (title, headers, rows) dataset to CSV / Excel / PDF responses."""
import csv
import io

from django.http import HttpResponse


def _slug(title: str) -> str:
    return title.lower().replace(" ", "_")


def export_csv(title, headers, rows):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{_slug(title)}.csv"'
    writer = csv.writer(response)
    writer.writerow(headers)
    writer.writerows(rows)
    return response


def export_excel(title, headers, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_fill = PatternFill("solid", fgColor="0D6EFD")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        ws.append(row)

    # Auto-size columns based on content length.
    for col, header in enumerate(headers, start=1):
        width = max(len(str(header)), *(len(str(r[col - 1])) for r in rows)) if rows else len(str(header))
        ws.column_dimensions[get_column_letter(col)].width = min(width + 2, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{_slug(title)}.xlsx"'
    return response


def export_pdf(title, headers, rows):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            title=title, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles["Title"]), Spacer(1, 12)]

    data = [headers] + [[str(c) for c in row] for row in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0D6EFD")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{_slug(title)}.pdf"'
    return response


EXPORTERS = {"csv": export_csv, "excel": export_excel, "pdf": export_pdf}
